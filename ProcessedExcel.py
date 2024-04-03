from datetime import datetime

import openpyxl

from openpyxl.utils import column_index_from_string

from dateutil.relativedelta import relativedelta

class Excel:

    def __init__(self, ExcelFile, Sheet, ExcelWechatName, ExcelWechatSentMessage, ExcelSentDate, ExcelAddDate):
        self.ExcelFileSaveLocation = ExcelFile
        self.ExcelFile = openpyxl.load_workbook(ExcelFile)
        self.ExcelFileSheet = self.ExcelFile[Sheet]
        self.ExcelWechatName = ExcelWechatName
        self.ExcelWechatSentMessage = ExcelWechatSentMessage
        self.ExcelSentDate = ExcelSentDate
        self.ExcelAddDate = ExcelAddDate
        self.ProcessedList = []
        self.Information = []

    def FilterExcel(self):

        self.ProcessedList.clear()

        today = datetime.now().date()


        for cell in self.ExcelFileSheet[self.ExcelSentDate]:

            try:

                cell_date = cell.value.date()

            except AttributeError:

                continue

            if cell_date == today:
                self.ProcessedList.append(cell.row)

    def FilterExcel(self):

        self.ProcessedList.clear()

        today = datetime.now().date()

        for cell in self.ExcelFileSheet[self.ExcelSentDate]:

            try:

                cell_date = cell.value.date()

            except AttributeError:

                continue

            if cell_date == today:
                self.ProcessedList.append(cell.row)

    def GetInformation(self):
        self.Information.clear()

        for row in self.ProcessedList:
            WechatName = self.ExcelFileSheet.cell(row=int(row), column=column_index_from_string(self.ExcelWechatName)).value
            WechatSentMessage = self.ExcelFileSheet.cell(row=int(row),
                                        column=column_index_from_string(self.ExcelWechatSentMessage)).value

            self.Information.append([WechatName, WechatSentMessage])

    def AddDate(self, AddRow):


        Addyear, Addmonth, Addday = self.ExcelFileSheet.cell(row=int(AddRow),
                                            column=column_index_from_string(self.ExcelAddDate)).value.split('-')

        cell_date = self.ExcelFileSheet.cell(row=int(AddRow), column=column_index_from_string(self.ExcelSentDate)).value.date()

        cell_date = cell_date + relativedelta(years=int(Addyear), months=int(Addmonth), days=int(Addday))

        self.ExcelFileSheet.cell(row=int(AddRow), column=column_index_from_string(self.ExcelSentDate)).value = cell_date

        self.ExcelFile.save(filename=str(self.ExcelFileSaveLocation))




